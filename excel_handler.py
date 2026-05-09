"""
excel_handler.py
================

Two responsibilities for the bulk-load workflow:

  1. write_template(decisions, screen, ...) → bytes
     Build an Excel workbook the user downloads, fills offline, and
     re-uploads. One sheet ("Data"), one column per field marked as
     "from Excel" on the review page. Header row uses the field NAME
     (e.g. POOLGROUPCODE) — that's the canonical key that matches what
     the screen uses internally and what our plan compiler refers to.

     **Per-column formatting is type-aware**:
       - DATE         → cells formatted YYYY-MM-DD
       - CHECKBOX     → list-validation dropdown ("Yes" / "No")
       - DROPDOWN /
         RADIO       → list-validation dropdown of the field's parsed options
       - NUMBER       → number format honouring precision
       - LOV-bound   → free text (user types the value to match)
       - Other (text) → free text

  2. read_uploaded(path) → list[dict[str, Any]]
     Parse a filled XLSX into a list of row dicts keyed by header.
     Trims trailing empty rows. Coerces simple types (numbers, dates) to
     reasonable string forms so the materialiser can drop them into "Enter
     <value>" steps without per-cell formatting logic.

We deliberately don't read the field LABEL or any contextual hint from
the spreadsheet — only the column headers (= field NAMES) matter. Users
can rearrange columns or add comment rows; we look up by header.
"""

from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SHEET_NAME = "Data"           # legacy name for the master sheet (single-sheet templates)
MASTER_SHEET_NAME = "Master"  # new name when multi-sheet templates have grids

# Number of data-row cells to pre-format / pre-validate. Excel still applies
# the validation if the user adds rows beyond this, but pre-applying ensures
# the dropdown arrow appears immediately on row 3.
DATA_ROW_RANGE = 200


def write_template(
    decisions: list[dict],
    screen: dict,
    *,
    placeholder_rows: int = 3,
) -> bytes:
    """Multi-sheet XLSX template:
      • Master sheet: master-block fields marked 'from Excel'.
      • One sheet per editable grid block: the grid's editable fields plus
        a `MASTER_KEY` column that links each grid row back to a master row.
        Read-only-only grids (e.g. FST_HIST) are skipped.

    A single-master, no-grid screen still produces a single sheet.
    """
    excel_decisions = [d for d in decisions if d.get("mode") == "excel"]
    field_lookup = {(f["block_name"], f["name"]): f for f in screen["fields"]}

    # Identify editable grids that should get their own sheet.
    blocks = screen.get("blocks") or []
    editable_grids: list[dict] = []
    for b in blocks:
        if not b.get("is_grid"):
            continue
        gfields = [f for f in screen["fields"] if f["block_name"] == b["name"]]
        if not gfields:
            continue
        if all(f.get("readonly") for f in gfields):
            continue  # read-only-only grid → omit from template
        editable_grids.append(b)

    # Master sheet should only carry non-grid fields. Grid fields are
    # already covered by their per-grid sheets below; including them on
    # the master sheet too would confuse users into filling them twice.
    grid_block_names = {b["name"] for b in editable_grids}
    master_decisions = [d for d in excel_decisions
                        if d.get("block_name") not in grid_block_names]

    # Empty case: no master decisions AND no editable grids.
    if not master_decisions and not editable_grids:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws["A1"] = "No fields marked 'from Excel' yet. Go back to the Review page."
        return _to_bytes(wb)

    wb = Workbook()
    # The default sheet becomes the master.
    master_ws = wb.active
    master_ws.title = MASTER_SHEET_NAME if editable_grids else SHEET_NAME

    # Custom in-screen buttons (Submit / Calculation / etc.) get per-row
    # Yes/No columns appended to the master sheet so each row decides
    # independently whether the runner should click them. Standard FLEXCUBE
    # toolbar buttons (`is_custom=False`) aren't user-toggleable.
    custom_buttons = [b for b in (screen.get("buttons") or []) if b.get("is_custom")]
    button_decisions_for_sheet = [
        {"block_name": "_buttons", "field_name": _button_column_name(b),
         "mode": "excel", "value": _button_column_name(b)}
        for b in custom_buttons
    ]
    button_field_lookup = {
        ("_buttons", _button_column_name(b)): {
            "block_name": "_buttons",
            "name":  _button_column_name(b),
            "label": (b.get("label") or b.get("name")),
            "datatype": "BUTTON_PRESS",
        }
        for b in custom_buttons
    }
    combined_field_lookup = {**field_lookup, **button_field_lookup}

    if master_decisions or button_decisions_for_sheet:
        _populate_sheet(
            master_ws,
            master_decisions + button_decisions_for_sheet,
            combined_field_lookup,
            placeholder_rows,
        )
    else:
        master_ws["A1"] = "(No master-block fields are 'from Excel'.)"

    # One sheet per editable grid.
    for grid in editable_grids:
        gfields = [f for f in screen["fields"]
                   if f["block_name"] == grid["name"] and not f.get("readonly")]
        ws = wb.create_sheet(_safe_sheet_name(grid["name"]))
        # Synthesize "decisions" for the grid sheet — every editable grid
        # field becomes a column. Plus a MASTER_KEY column up front.
        master_key_field = {
            "block_name": "_meta", "name": "MASTER_KEY",
            "label": "Master record key",
            "datatype": "VARCHAR2",
        }
        synthetic_decisions = (
            [{"block_name": "_meta", "field_name": "MASTER_KEY", "mode": "excel", "value": "MASTER_KEY"}]
            + [{"block_name": grid["name"], "field_name": f["name"], "mode": "excel", "value": f["name"]}
               for f in gfields]
        )
        synthetic_lookup = dict(field_lookup)
        synthetic_lookup[("_meta", "MASTER_KEY")] = master_key_field
        _populate_sheet(ws, synthetic_decisions, synthetic_lookup, placeholder_rows,
                        intro=f"Grid: {grid.get('label') or grid['name']} — "
                              "MASTER_KEY links each row to the matching master.")

    return _to_bytes(wb)


def _populate_sheet(
    ws,
    excel_decisions: list[dict],
    field_lookup: dict,
    placeholder_rows: int,
    intro: str | None = None,
) -> None:
    name_font   = Font(bold=True, color="FFFFFF")
    label_font  = Font(italic=True, color="666666", size=10)
    header_fill = PatternFill("solid", fgColor="2A3142")
    centered    = Alignment(horizontal="left", vertical="center")

    for idx, d in enumerate(excel_decisions, start=1):
        col_letter = get_column_letter(idx)
        field = field_lookup.get((d.get("block_name"), d["field_name"])) or {}
        name = d["field_name"]
        datatype = (field.get("datatype") or "").upper()
        hint = _format_hint(field, datatype)

        cell = ws.cell(row=1, column=idx, value=name)
        cell.font = name_font
        cell.fill = header_fill
        cell.alignment = centered

        cell = ws.cell(row=2, column=idx, value=hint)
        cell.font = label_font
        cell.alignment = centered

        _apply_column_format(ws, idx, col_letter, field, datatype)
        ws.column_dimensions[col_letter].width = max(len(name), len(hint), 16) + 4

    ws.freeze_panes = "A3"

    if intro:
        # Optional intro note above the data rows; doesn't affect parsing
        # because reader skips row 2 by design.
        last_col = get_column_letter(len(excel_decisions))
        ws.merge_cells(f"A{3 + placeholder_rows}:{last_col}{3 + placeholder_rows}")
        note = ws.cell(row=3 + placeholder_rows, column=1, value=intro)
        note.font = Font(italic=True, color="888888", size=10)


def _button_column_name(btn: dict) -> str:
    """Excel column header for a custom in-screen button. Prefix `Press_`
    is load-bearing — `read_uploaded_full` keys on it to extract per-row
    button decisions from each parsed row."""
    return f"Press_{btn.get('name')}"


def _safe_sheet_name(name: str) -> str:
    # Excel sheet names: max 31 chars, no `:\\/?*[]`
    cleaned = "".join("_" if c in r":\/?*[]" else c for c in name)
    return cleaned[:31] or "Sheet"


def read_uploaded(path: Path) -> list[dict[str, Any]]:
    """Parse the master sheet of a filled XLSX into row dicts. Kept for
    backwards compatibility with single-sheet templates."""
    return read_uploaded_full(path)["_master"]


def read_uploaded_full(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Parse every sheet. Returns:
        {"_master": [...master rows], "<grid_block_name>": [...rows], ...}

    Tolerates either a Master-named sheet (multi-sheet template) or a
    Data-named single sheet (legacy)."""
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    out: dict[str, list[dict[str, Any]]] = {"_master": []}

    # Pick the master sheet
    if MASTER_SHEET_NAME in wb.sheetnames:
        master_ws = wb[MASTER_SHEET_NAME]
    elif SHEET_NAME in wb.sheetnames:
        master_ws = wb[SHEET_NAME]
    else:
        master_ws = wb.active

    out["_master"] = _read_sheet_rows(master_ws)

    # Every other sheet is a grid sheet — keyed by its title.
    for sheet_name in wb.sheetnames:
        if sheet_name in (MASTER_SHEET_NAME, SHEET_NAME):
            continue
        out[sheet_name] = _read_sheet_rows(wb[sheet_name])
    return out


def _read_sheet_rows(ws) -> list[dict[str, Any]]:
    headers: list[str] = []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        return []
    for cell in first:
        headers.append(str(cell).strip() if cell is not None else "")
    try:
        next(rows_iter)  # skip the "(label)" hint row
    except StopIteration:
        return []
    out: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row = {h: _coerce(raw_row[i] if i < len(raw_row) else None)
               for i, h in enumerate(headers) if h}
        if not any(v not in (None, "") for v in row.values()):
            continue
        out.append(row)
    return out


# ---------------------------------------------------------------------------
# Per-column format helpers
# ---------------------------------------------------------------------------

CHECKBOX_TRUTHY = {"YES", "Y", "TRUE", "1", "TICK", "TICKED", "CHECKED"}


def _format_hint(field: dict, datatype: str) -> str:
    """Human-readable cue printed in row 2 of each column."""
    name = field.get("name") or ""
    label = field.get("label") or name
    if name == "MASTER_KEY":
        return ("BLANK = apply to every master record. "
                "Or fill 1/2/3/… to link this row to one specific master.")
    if datatype == "BUTTON_PRESS":
        return f"({label}) — Yes to click this button for this row, No / blank to skip"
    if datatype == "DATE":
        return f"({label}) — date YYYY-MM-DD"
    if datatype == "CHECKBOX":
        return f"({label}) — Yes (or blank to leave default)"
    if datatype in ("DROPDOWN", "RADIO"):
        opts = field.get("options") or []
        if opts:
            preview = " | ".join(str(o.get("value", "")) for o in opts[:4])
            if len(opts) > 4:
                preview += " | …"
            return f"({label}) — one of: {preview}"
        return f"({label}) — option value"
    if datatype == "NUMBER":
        precision = field.get("precision")
        suffix = f" ({precision} d.p.)" if precision else ""
        return f"({label}) — number{suffix}"
    if field.get("lov"):
        return f"({label}) — value to match (LOV {field['lov']})"
    if datatype in ("VARCHAR2", "VARCHAR", "TEXT", "CHAR"):
        length = field.get("length")
        return f"({label}) — text" + (f" (max {length})" if length else "")
    return f"({label})"


def _apply_column_format(
    ws,
    col_idx: int,
    col_letter: str,
    field: dict,
    datatype: str,
) -> None:
    """Apply Excel cell formatting + data validation for the given field's
    type. Operates on rows 3..(2 + DATA_ROW_RANGE) — enough that the user
    sees the formatting immediately when they start typing."""
    rng_str = f"{col_letter}3:{col_letter}{2 + DATA_ROW_RANGE}"

    # MASTER_KEY column on grid sheets: integer master-row number.
    if field.get("name") == "MASTER_KEY":
        for r in range(3, 3 + DATA_ROW_RANGE):
            ws.cell(row=r, column=col_idx).number_format = "0"
        return

    # Custom-button press toggle: Yes/No dropdown, identical mechanics to
    # the CHECKBOX validator. Read-back uses the same Yes/Y/True/1 truthy
    # set so the bulk composer can map the cell to a click decision.
    if datatype == "BUTTON_PRESS":
        dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="warning",
            error="Use Yes or No (or leave blank to skip).",
        )
        dv.showDropDown = False
        dv.add(rng_str)
        ws.add_data_validation(dv)
        return

    # LOV-bound fields are free text. Don't add validation; let the user
    # type the value-to-match.
    if field.get("lov"):
        return

    if datatype == "DATE":
        for r in range(3, 3 + DATA_ROW_RANGE):
            ws.cell(row=r, column=col_idx).number_format = "YYYY-MM-DD"
        return

    if datatype == "CHECKBOX":
        # Excel doesn't have native cell-level checkboxes without ActiveX.
        # The standard pattern is a Yes/No data-validation dropdown.
        dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="warning",
            error="Use Yes or No (or leave blank).",
        )
        # showDropDown is inverted in OOXML — `False` makes the arrow visible.
        dv.showDropDown = False
        dv.add(rng_str)
        ws.add_data_validation(dv)
        return

    if datatype in ("DROPDOWN", "RADIO"):
        opts = field.get("options") or []
        if not opts:
            return
        # OOXML literal-list formula1 is capped at ~250 chars including
        # surrounding quotes. Skip data-validation for unusually-long sets;
        # the user can still type a value, and the materialiser still maps
        # it correctly.
        values = [str(o.get("value", "")).replace('"', "'") for o in opts]
        joined = ",".join(v for v in values if v)
        if 0 < len(joined) <= 250:
            dv = DataValidation(
                type="list",
                formula1=f'"{joined}"',
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="warning",
            )
            dv.showDropDown = False
            dv.add(rng_str)
            ws.add_data_validation(dv)
        return

    if datatype == "NUMBER":
        precision = field.get("precision") or 0
        fmt = "0" if not precision else "0." + ("0" * min(precision, 6))
        for r in range(3, 3 + DATA_ROW_RANGE):
            ws.cell(row=r, column=col_idx).number_format = fmt
        return

    # VARCHAR / TEXT / CHAR / unknown → no special formatting (free text).


# ---------------------------------------------------------------------------
# Cell value coercion
# ---------------------------------------------------------------------------

def _coerce(v: Any) -> str | None:
    """Coerce a cell value into the string the materialiser drops into
    `Enter <value>` instructions. The materialiser itself does the
    type-aware mode mapping (e.g. 'Yes' → tick); here we only normalise."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, bool):
        # openpyxl can occasionally surface BOOLEANs from formulas; map to
        # the same forms our checkbox materialiser recognises.
        return "Yes" if v else "No"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
